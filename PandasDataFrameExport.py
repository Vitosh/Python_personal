import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook

#Page setup
wb = Workbook()
ws0 = wb.active
ws0.title = "SomeName"
ws0.sheet_properties.tabColor = "1072BA"
wb.save('Example.xlsx')
wb.close()

#Insert pandas series on existing worksheet
path = r".\Example.xlsx"
book = load_workbook('Example.xlsx')
writer = pd.ExcelWriter(path, engine = 'openpyxl')
writer.book = book
ws0 = book.active
r1y = pd.Series([1,3,5,13,6,8])
df1y = pd.DataFrame(r1y)

df1y.to_excel(writer, "SomeName", header=False, startcol = 0, startrow = 1)
writer.save()
writer.close()
