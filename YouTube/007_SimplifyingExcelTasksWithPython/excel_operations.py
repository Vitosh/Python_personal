import pandas as pd
from openpyxl import load_workbook


def read_excel(file_path):
    df = pd.read_excel(file_path)
    print("Original Data:")
    print(df)
    return df

def manipulate_data(df):
    df['Total'] = df['Quantity'] * df['UnitPrice']
    df['TotalWithVAT'] = 0
    print("Manipulated Data:")
    print(df)
    return df

def write_to_excel_with_formulas(df, output_file_path):
    df.to_excel(output_file_path, index=False)
    workbook = load_workbook(output_file_path)
    worksheet = workbook.active

    for row in range(2, len(df)+2):
        cell = worksheet.cell(row=row, column=4)
        cell.value = f'=C{row}*1.2'

    workbook.save(output_file_path)
    print(f"Data with formulas has been written to {output_file_path}")

if __name__ == "__main__":
    input_file_path = 'input.xlsx'
    output_file_path = 'output.xlsx'

    data = read_excel(input_file_path)
    manipulated_data = manipulate_data(data)
    write_to_excel_with_formulas(manipulated_data, output_file_path)